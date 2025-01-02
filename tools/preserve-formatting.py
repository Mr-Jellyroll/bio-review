import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter

def list_xlsx_files():
    """
    List all .xlsx files in the current directory.
    """
    return [file for file in os.listdir() if file.endswith(".xlsx")]

def save_formatting_and_widths(original_file):
    """
    Save formatting and column widths from the original .xlsx file.
    Returns a dictionary of formatting and a dictionary of column widths.
    """
    wb = load_workbook(original_file)
    ws = wb.active
    formatting = {}
    column_widths = {}

    for col in ws.column_dimensions:
        column_widths[col] = ws.column_dimensions[col].width

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                formatting[cell.coordinate] = {
                    "font": cell.font.copy(),
                    "fill": cell.fill.copy(),
                    "alignment": cell.alignment.copy(),
                    "border": cell.border.copy(),
                }
    return formatting, column_widths

def apply_formatting_and_widths(ws, formatting, column_widths):
    """
    Apply saved formatting and column widths to the worksheet.
    """
    for coord, fmt in formatting.items():
        cell = ws[coord]
        cell.font = fmt["font"]
        cell.fill = fmt["fill"]
        cell.alignment = fmt["alignment"]
        cell.border = fmt["border"]

    for col, width in column_widths.items():
        if width is not None:
            ws.column_dimensions[col].width = width

def convert_xlsx_to_csv(input_file, output_csv):
    """
    Convert an .xlsx file to .csv.
    """
    data = pd.read_excel(input_file, engine="openpyxl", keep_default_na=False)
    data.to_csv(output_csv, index=False)
    print(f"Converted {input_file} to {output_csv}")

def convert_csv_to_xlsx(input_csv, output_xlsx, formatting, column_widths):
    """
    Convert a .csv file to .xlsx and apply formatting and column widths.
    """
    # Read CSV into DataFrame
    data = pd.read_csv(input_csv)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write data into worksheet
    for r_idx, row in enumerate(data.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply formatting and column widths
    apply_formatting_and_widths(ws, formatting, column_widths)

    # Save the workbook
    wb.save(output_xlsx)
    print(f"Converted {input_csv} to {output_xlsx} with formatting and column widths applied")

if __name__ == "__main__":
    # Step 1: List all .xlsx files
    xlsx_files = list_xlsx_files()
    if not xlsx_files:
        print("No .xlsx files found in the current directory.")
        exit()

    print("Available .xlsx files:")
    for idx, file in enumerate(xlsx_files, start=1):
        print(f"{idx}: {file}")

    # Step 2: Ask the user to select a file
    try:
        choice = int(input("Enter the number of the file to convert: ")) - 1
        if choice < 0 or choice >= len(xlsx_files):
            print("Invalid choice. Exiting.")
            exit()
        original_xlsx = xlsx_files[choice]
    except ValueError:
        print("Invalid input. Please enter a number. Exiting.")
        exit()

    intermediate_csv = "intermediate.csv"
    final_xlsx = f"final_{original_xlsx}"  # Create a new name for the final file

    # Step 3: Save formatting and column widths from the selected .xlsx
    formatting, column_widths = save_formatting_and_widths(original_xlsx)

    # Step 4: Convert selected .xlsx to .csv
    convert_xlsx_to_csv(original_xlsx, intermediate_csv)

    # Step 5: Convert .csv back to .xlsx and apply formatting and column widths
    convert_csv_to_xlsx(intermediate_csv, final_xlsx, formatting, column_widths)

    print(f"Process completed. Final file: {final_xlsx}")
